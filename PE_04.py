import os
import pefile
import openpyxl as pyxl
import pprint as ppr
import re
from openpyxl.styles import PatternFill, Color

class Excel_stu:
    # 생성자
    def __init__(self):
        self.workBook = pyxl.Workbook()  # 객체 생성
        self.workSheet = None
        self.color_list = ['ffc3a0', '86cece', 'e5e545', 'f5ebd1', 'cd6090', 'b8d9d0', 'ffebe1']
    def dos_header_excel_save(self):
        self.workBook.save("C:/Users/sleep/Desktop/Reversing/dos_header.xlsx")


        '''
        C3 ~ R3
        e_cblp': {'Offset': '02'
        '''
class PE_DOS_HEADER:
    def __init__(self):
        self.left_re  = re.compile('\[')
        self.right_re = re.compile('\]')
        self.dos_header_dict = dict() # type of dictionary
        self.dos_header_element = \
"""WORD e_magic
WORD e_cblp
WORD e_cp
WORD e_crlc
WORD e_cparhdr
WORD e_minalloc
WORD e_maxalloc
WORD e_ss
WORD e_sp
WORD e_csum
WORD e_ip
WORD e_cs
WORD e_lfarlc
WORD e_ovno
WORD e_res[4]
WORD e_oemid
WORD e_oeminfo
WORD e_res2[10]
DWORD e_lfanew
"""
    def check(self):
        t_list = self.dos_header_element.split(sep='\n')
        t_list = [i.split(' ') for i in t_list]
        for i in t_list:
            if i[0] == 'WORD':
                l_value = self.left_re.search(i[1])
                if l_value:
                    r_value = self.right_re.search(i[1])
                    l_index, r_index = l_value.span()[1], r_value.span()[0]
                    t_size = int(i[1][l_index:r_index])
                    k_index = l_value.span()[0]
                    self.dos_header_dict[i[1][:k_index]] = {"SIZE": t_size * 2, "Offset":None}
                else: # False
                    self.dos_header_dict[i[1]] = {"SIZE": 2, "Offset": None}

            elif i[0] == 'DWORD':
                l_value = self.left_re.search(i[1])
                if l_value:
                    r_value = self.right_re.search(i[1])
                    l_index, r_index = l_value.span()[1], r_value.span()[0]
                    t_size = int(i[1][l_index:r_index])
                    k_index = l_value.span()[0]
                    self.dos_header_dict[i[1][:k_index]] = {"SIZE": t_size * 4, "Offset": None}
                else:
                    self.dos_header_dict[i[1]] = {"SIZE": 4, "Offset": None}
        #ppr.pprint (self.dos_header_dict)

class PE_stu(Excel_stu, PE_DOS_HEADER):
    def __init__(self):
        # EXCEL
        Excel_stu.__init__(self) # 부모 클래스 초기화
        PE_DOS_HEADER.__init__(self)

        self.target   = 'C:/Users/sleep/Desktop/Reversing/putty.exe'
        self.peObject = pefile.PE(self.target)
        self.DOS_HEADER_reObject = re.compile(pattern="^e")

    # DOS_HEADER 64_________________________________________________
    def dos_header(self):
        dos_info = self.peObject.DOS_HEADER.dump_dict().items()
        #ppr.pprint (dos_info)
        for k, v in dos_info:
            if self.DOS_HEADER_reObject.search(k):
                self.dos_header_dict[k]["Offset"] = "%02x"%(v['Offset'])
        ppr.pprint (self.dos_header_dict)
        """
        e_lfanew, e_magic
        """
    def dos_header_excel(self):
        self.workSheet = self.workBook.create_sheet('DOS_HEADER')
        INDX = ['C', 3] # type of tuple
        with open(self.target,"rb") as f:
            for _ in range(64):
                word = f.read(1)
                s = "{0:02x}".format(ord(word))
                self.workSheet[INDX[0] + str(INDX[1])] = s
                if INDX[0] == 'R':
                    INDX[0] = 'C'
                    INDX[1] += 1
                else: # INDX[0] != 'R'
                    INDX[0] = chr(ord(INDX[0]) + 1)

    def dos_header_excel_color(self):
        t_dos_excel = pyxl.load_workbook("C:/Users/sleep/Desktop/Reversing/dos_header.xlsx")
        ws = t_dos_excel.sheetnames
        ws = t_dos_excel[ws[1]]
        INDX = ['C', 3]
        c = 0
        for v in self.dos_header_dict.values():
            indx, size = int(v['Offset'], 16), v['SIZE']
            for a in range(0, size):
                t_data, t_mov = (indx + a)//16, (indx + a)%16
                if t_data   == 0:
                    ws[chr(ord(INDX[0])+t_mov) + str(INDX[1] + t_data)].fill = PatternFill(
                        patternType='solid', fgColor=Color(self.color_list[c]))
                elif t_data == 1:
                    ws[chr(ord(INDX[0])+t_mov) + str(INDX[1] + t_data)].fill = PatternFill(
                        patternType='solid', fgColor=Color(self.color_list[c]))
                elif t_data == 2:
                    ws[chr(ord(INDX[0])+t_mov) + str(INDX[1] + t_data)].fill = PatternFill(
                        patternType='solid', fgColor=Color(self.color_list[c]))
                else:
                    ws[chr(ord(INDX[0])+t_mov) + str(INDX[1] + t_data)].fill = PatternFill(
                        patternType='solid', fgColor=Color(self.color_list[c]))
            c = (c + 1) % len(self.color_list)
        t_dos_excel.save("C:/Users/sleep/Desktop/Reversing/dos_header.xlsx")
        print ("success save")
def main():
    myPE = PE_stu()
    myPE.check()
    myPE.dos_header()
    myPE.dos_header_excel()
    myPE.dos_header_excel_color()
if __name__ == "__main__":
    main()