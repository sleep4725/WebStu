from selenium import webdriver
import time
from bs4 import BeautifulSoup
from tinydb import TinyDB
from openpyxl import Workbook, load_workbook
import re
import pprint as ppr
from openpyxl.styles import PatternFill, Color
# =============================
"""
//*[@id="smGiList"]/div[3]/ul/li[1]/span
//*[@id="smGiList"]/div[3]/ul/li[2]/span
"""
class JobKorea:
    def __init__(self, s):
        # instance 변수 셋팅
        # Chrome driver open
        self.browser = webdriver.Chrome("C:\\Users\\sleep\\Desktop\\getDriveFirfox\\chromedriver.exe")
        self.searchData = s
        self.Data = []
        self.pageValue = 0x0

        # re-
        self.filter_todayEnd = re.compile(pattern='오늘마감')
        self.filter_always = re.compile(pattern='상시모집')
        self.filter_date = re.compile('(\d{1,2})/(\d{1,2})')
        # ===================================================

        # excel-
        self.wb = Workbook() # 객체 생성, workBook
        self.ws = None # workSheet
        self.loadWkbook = None
        # ===================================================
        self.execel_dict = {} # type of dictionary

    # instance method - 1
    def URLrequest(self):
        self.browser.get("http://www.jobkorea.co.kr/Search/")
        # 'Python' 검색
        self.browser.find_element_by_id('stext').send_keys(self.searchData)
        self.browser.find_element_by_xpath('//*[@id="common_search_btn"]').click()

        # 조건 : 근무 경력 - 신입
        self.browser.find_element_by_xpath('//*[@id="search_wrap"]/dl[5]/dd/div[1]/ul/li[1]/label/span/span').click()

        # 시간 : 3초 - 반드시 줄 것 아니면 값이 제대로 들어가지 않음
        time.sleep(3)
        self.PageURLrequest()

    # instance method - 2
    def PageURLrequest(self):
        p = 1
        # page
        while True:
            try:
                self.browser.find_element_by_xpath('//*[@id="smGiList"]/div[3]/ul/li[{}]'.format(p)).click()
            except:
                self.pageValue = p - 1
                break
            else: # 성공
                # 시간 : 3초 - 반드시 줄 것 아니면 값이 제대로 들어가지 않음
                time.sleep(3)
                html = self.browser.page_source  # HTML.text
                bs_object = BeautifulSoup(html, "html.parser")
                title = bs_object.find_all('span', {'class': 'corpName'})
                infov = bs_object.find_all('span', {'class': 'detailInfo'})
                print("{}:page =================".format(p))
                for i, t in enumerate(zip(title, infov)):
                    unit_data = {} # dictionary data

                    name = t[0].select_one('a')
                    tinf = t[1].select_one('p.gibInfo > a')
                    tmpv = t[1].select_one('p.gibDesc > a > em')
                    tinf = re.sub('[\n\r\t ]', '', tinf.text) # \n과 \r, \t, , , 를 제거한다.
                    tinf = tinf.split(',') # list 화

                    # name : 회사 이름
                    print("{0:02d} => {1:s}, {2}".format(i + 1, name.string, tinf))
                    unit_data[name.string] = {'info':tinf,
                                              'date':None,
                                              'page':self.searchData + '_' + str(p)}


                    if str(tmpv.string).startswith('~'):
                        t = self.filter_date.search(str(tmpv.string)[1:])
                        tIndx = t.span()
                        print (str(tmpv.string)[1:][tIndx[0]:tIndx[1]])
                        unit_data[name.string]['date'] = str(tmpv.string)[1:][tIndx[0]:tIndx[1]]
                    else:
                        # 오늘 마감
                        if self.filter_todayEnd.search(str(tmpv.string)):
                            print ("오늘마감")
                            unit_data[name.string]['date'] = '오늘마감'
                        # 상시 모집
                        elif self.filter_always.search(str(tmpv.string)):
                            print ("상시모집")
                            unit_data[name.string]['date'] = '상시모집'
                        else:
                            print (str(tmpv.string))
                            unit_data[name.string]['date'] = str(tmpv.string)
                    self.Data.append(unit_data)
                print("========================")
                p = p+1 # page up

    # instance method - 3 sheet 생성
    def createSheet(self):
        for i in range(1, self.pageValue+1):
            self.wb.create_sheet(self.searchData + '_' + str(i))
        self.wb.save('C:\\Users\\sleep\\Desktop\\jobKorea_{}.xlsx'.format(self.searchData))
        print ("save success ...")

    # instance method - 4
    def excelPageSetting(self):
        # 엑셀 _ controller
        for p in range(1, self.pageValue+1):
            self.execel_dict[self.searchData + '_' + str(p)] = \
                        {"회사이름": [ord('B'), 3],
                         "정보": [ord('B'), 4],
                         "날짜": [ord('B'), 5],
                         "번호": [ord('A'), 3],
                         "count": 1}

    # instance method - 5 : loadWorkbook
    def xlsxWrite(self):
        # 엑셀 파일 load
        self.loadWkbook = load_workbook('C:\\Users\\sleep\\Desktop\\jobKorea_{}.xlsx'.format(self.searchData))
        # self.loadWkbook.sheetnames # ['Sheet', 'python_1', 'python_2', 'python_3', 'python_4', 'python_5', 'python_6']
        for i in range(0, len(self.Data)):
            for k, v in self.Data[i].items():
                tempJump = 0x0
                # 엑셀 내용 저장
                WkSheet = self.loadWkbook[v['page']] # worksheet load
                # 번호) ---------------------------------------------
                inxCha = self.execel_dict[v['page']]['번호'][0]
                inxInt = self.execel_dict[v['page']]['번호'][1]
                WkSheet[chr(inxCha) + str(inxInt)] = self.execel_dict[v['page']]['count']

                # # 회사 이름) ---------------------------------------
                inxCha = self.execel_dict[v['page']]['회사이름'][0]
                inxInt = self.execel_dict[v['page']]['회사이름'][1]
                WkSheet[chr(inxCha) + str(inxInt)] = "회사이름" # - 회사이름
                # 셀 너비 조정
                # WkSheet.column_dimensions[chr(inxCha) + str(inxInt)].width = len("회사이름")
                inxCha = inxCha + 1
                WkSheet[chr(inxCha) + str(inxInt)] = k # - 구글

                # 요구 기술) -----------------------------------------
                inxCha = self.execel_dict[v['page']]['정보'][0] # "B"
                inxInt = self.execel_dict[v['page']]['정보'][1] #  4
                WkSheet[chr(inxCha) + str(inxInt)] = "정보"  # - 정보
                tmpWidth = 0x0
                inxCha = inxCha + 1
                for j in range(0, len(v['info'])):
                    if inxCha > ord('Z'): # 값이 넘어 가는 경우
                        inxCha = ord('C')
                        tempJump += 1
                        inxInt += 1
                    WkSheet[chr(inxCha) + str(inxInt)] = v['info'][j]
                    # 셀 병합 ==========================================
                    front = chr(inxCha) + str(inxInt)
                    if inxCha+1 > ord('Z'):
                        rear = 'AA' + str(inxInt)
                    else: # inxCha+1 <= ord('Z')
                        rear = chr(inxCha+1) + str(inxInt)
                    WkSheet.merge_cells(front+':'+rear)
                    # =================================================
                    inxCha = inxCha + 2
                    # =================================================
                    # 열 너비 조정
                    # if len(v['info'][j]) > tmpWidth:
                    #     WkSheet.column_dimensions[chr(inxCha)].width = len(v['info'][j])
                    #     tmpWidth = len(v['info'][j])

                    # 셀 너비 조정
                    # WkSheet.column_dimensions[chr(inxCha) + str(inxInt)].width = len(v['info'][j])

                if self.execel_dict[v['page']]['정보'][1] != inxInt:
                    self.execel_dict[v['page']]['날짜'][1] = inxInt + 1
                # 날짜) ---------------------------------------------
                inxCha = self.execel_dict[v['page']]['날짜'][0]
                inxInt = self.execel_dict[v['page']]['날짜'][1]
                WkSheet[chr(inxCha) + str(inxInt)] = "마감날짜"
                inxCha = inxCha + 1
                WkSheet[chr(inxCha) + str(inxInt)] = v['date']
                if v['date'] == "오늘마감":
                    WkSheet[chr(inxCha) + str(inxInt)].fill = \
                        PatternFill(patternType='solid', fgColor=Color('ff547d'))
                # ===========================================================
                # 4씩 값을 증가
                self.execel_dict[v['page']]['번호'][1] += (4 + tempJump)
                self.execel_dict[v['page']]['회사이름'][1] += (4 + tempJump)
                self.execel_dict[v['page']]['정보'][1] += (4 + tempJump)
                self.execel_dict[v['page']]['날짜'][1] += 4
                self.execel_dict[v['page']]['count'] += 1

                # self.loadWkbook.save('C:\\Users\\sleep\\Desktop\\jobKorea_{}.xlsx'.format(self.searchData))
                # exit(1)
        self.loadWkbook.save('C:\\Users\\sleep\\Desktop\\jobKorea_{}.xlsx'.format(self.searchData))
    # 소멸자 -
    # def __del__(self):
    #     self.wb.save('C:\\Users\\sleep\\Desktop\\jobKorea_{}.xlsx'.format(self.searchData))

def main():
    jobNode = JobKorea('빅데이터') # 객체 생성
    jobNode.URLrequest() # instance method - 1
    jobNode.createSheet() # instance method - 3
    jobNode.excelPageSetting() # instance method - 4
    jobNode.xlsxWrite() # instance method - 5
    # ppr.pprint (jobNode.Data)
if __name__ == "__main__":
    main()