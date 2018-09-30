#========================================
import requests
from bs4 import BeautifulSoup
import pprint as ppr
import re
import openpyxl as opxl
from openpyxl.styles import PatternFill, Color
#========================================
class SaramInfo:
    # 생성자
    def __init__(self, url):
        self.target_url = url
        self.html = None
        self.bs_object = None   # BeautifulSoup
        self.total_corp_info = dict()
        self.workBook = opxl.Workbook()
        self.page_info = None

    # [+] func #1
    def RequestsSaramIn(self):
        p_ = 1
        while p_:
            tmp_parameter = {
                "searchword":"python",
                "page":p_,
            }
            self.total_corp_info[p_] = []
            self.html = requests.get(self.target_url, tmp_parameter)
            if self.html.ok and self.html.status_code == 200:
                print ("현재 페이지: {0:d} 처리 중...".format(p_))
                self.bs_object = BeautifulSoup(self.html.text, "html.parser")

                corp_info = self.bs_object.select("#recruit_info_list > ul > li > div")

                # 페이지 스타트 __________________________________
                # print("[page]: {0:d}".format(p_))

                # [회사 정보]
                if self.total_corp_info[p_] == []:
                    self.total_corp_info[p_] = list([{"회사이름":i.select('div > h2 > a > span')[0].string,
                                              "경력":i.select('p.terms_li > span:nth-of-type(1)')[0].string,
                                              "키워드":[k.string for k in i.select('p.keywordline > span > a')],
                                              "날짜":re.sub('[\r,\n, ]', '', str(i.select('p.txt > em')[0].string))} for i in corp_info])
                else:
                    # self.total_corp_info[p_] != None
                    self.total_corp_info[p_].append(
                        list([{"회사이름": i.select('div > h2 > a > span')[0].string,
                               "경력": i.select('p.terms_li > span:nth-of-type(1)')[0].string,
                               "키워드": [k.string for k in i.select('p.keywordline > span > a')],
                               "날짜": re.sub('[\r,\n, ]', '', str(i.select('p.txt > em')[0].string))} for i in corp_info])
                    )

                if self.total_corp_info[p_] == []:
                    break
                else:
                    # ppr.pprint (self.total_corp_info)
                    print("현재 페이지: {0:d} 종료".format(p_))
                    p_ += 1

        self.page_info = p_
        print ("현재 검색어 : {} 에 대한 전체 페이지 수 : {}".format(tmp_parameter["searchword"],
                                                       self.page_info))
    # [+] func #2
    def excelWrite(self):

        for i in range(1, self.page_info+1):
            # ========================================================
            # INDEX = ['B', 4] # type of tuple
            # LstIndex = ['C', 4]
            LstInfo = ['회사이름', '경력', '키워드', '날짜']
            # DataIndex = ['D', 4]
            index_1 = 0x0
            IndexList = ['C', [4,5,6,7], 5]
            #========================================================
            ws = self.workBook.create_sheet('page_{0:d}'.format(i))
            ws['B3'] = "번호"

            sub_cor_info = [d for d in self.total_corp_info[i]]
            for data in sub_cor_info:
                # 데이터 출력 후에 주석으로 막을 것
                # for j in range(len(LstInfo)):
                #     print ("{}:  {}".format(LstInfo[j], data[LstInfo[j]]))
                for k in range(0, len(IndexList[1])):
                    if LstInfo[k] == "회사이름":
                        ''' C + 4  : 회사 이름'''
                        ws[IndexList[0]+str(IndexList[1][k])] = LstInfo[k]
                        ''' D '''
                        tmpIndex = chr(ord(IndexList[0]) + 1)
                        ''' D + 4'''
                        ws[tmpIndex+str(IndexList[1][k])] = data[LstInfo[k]]
                    elif LstInfo[k] == "경력":
                        ws[IndexList[0] + str(IndexList[1][k])] = LstInfo[k]
                        tmpIndex = chr(ord(IndexList[0]) + 1)
                        ws[tmpIndex + str(IndexList[1][k])] = data[LstInfo[k]]
                        if data[LstInfo[k]] == "경력무관":
                            ws[tmpIndex + str(IndexList[1][k])].fill = PatternFill(patternType='solid', fgColor=Color('fac1e2'))

                    elif LstInfo[k] == "키워드":
                        ws[IndexList[0] + str(IndexList[1][k])] = LstInfo[k]
                        tmpIndex = chr(ord(IndexList[0]) + 1)
                        datav = ','.join(data[LstInfo[k]])
                        ws[tmpIndex + str(IndexList[1][k])] = datav

                    else: # "날짜"
                        ws[IndexList[0] + str(IndexList[1][k])] = LstInfo[k]
                        tmpIndex = chr(ord(IndexList[0]) + 1)
                        # print (tmpIndex + str(IndexList[1][k]))
                        ws[tmpIndex + str(IndexList[1][k])] = data[LstInfo[k]]

                for i_s in range(0, len(IndexList[1])):
                    IndexList[1][i_s] = IndexList[1][i_s] + IndexList[2]



                # # B4 : 번호 ----------------------------------------
                # # ws[INDEX[0] + str(INDEX[1])] = str(s[0] + 1)
                # C4 : 목록 - 회사 이름
                # ws[LstIndex[0] + str(LstIndex[1])] = LstInfo[LstInfo_index]
                # D4 : 데이터
                # ws[DataIndex[0] + str(DataIndex[1])] = data[LstInfo[LstInfo_index]]
                # #===================================================
                # LstIndex[1] += 1
                # LstInfo_index = (LstInfo_index + 1)%len(LstInfo)
                # DataIndex[1] += 1
                # # ===================================================
                # # C4 : 목록 - 경력
                # ws[LstIndex[0] + str(LstIndex[1])] = LstInfo[LstInfo_index]
                # ws[DataIndex[0] + str(DataIndex[1])] = data[LstInfo[LstInfo_index]]
                # # ===================================================
                # LstIndex[1] += 1
                # LstInfo_index = (LstInfo_index + 1) % len(LstInfo)
                # DataIndex[1] += 1
                # tmp = DataIndex
                # # ===================================================
                # # C4 : 목록 - 키워드
                # ws[LstIndex[0] + str(LstIndex[1])] = LstInfo[LstInfo_index]
                # # ws[DataIndex[0] + str(DataIndex[1])] = s[LstInfo[LstInfo_index]]
                # tmpList = [ l for l in data[LstInfo[LstInfo_index]]]
                # for k in tmpList:
                #     ws[tmp[0] + str(tmp[1])] = k
                #     tmp[0] = chr(ord(tmp[0]) + 1)
                # # ===================================================
                # LstIndex[1] += 1
                # LstInfo_index = (LstInfo_index + 1) % len(LstInfo)
                # DataIndex[1] += 1
                # # ===================================================
                # # C4 : 목록 - 날짜
                # ws[LstIndex[0] + str(LstIndex[1])] = LstInfo[LstInfo_index]
                # ws[DataIndex[0] + str(DataIndex[1])] = data[LstInfo[LstInfo_index]]
                # # ===================================================
                # LstIndex[1] += 1
                # LstInfo_index = (LstInfo_index + 1) % len(LstInfo)
                # DataIndex[1] += 1

    # 소멸자
    def __del__(self):
        print ("성공적으로 저장되었습니다. ... ")
        self.workBook.save("사람인_채용_관련_데이터.xlsx")

def main():
    sNode = SaramInfo("http://www.saramin.co.kr/zf_user/search")
    sNode.RequestsSaramIn()

    sNode.excelWrite()
if __name__ == "__main__":
    main()